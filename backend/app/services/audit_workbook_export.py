"""Multi-sheet auditor workbook (xlsx) for compliance reviews + STIG summary tables."""

from __future__ import annotations

import io
import json
from datetime import datetime
from typing import Any, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.services.metric_semantics import is_open_risk

STIG_RULE_ROWS_CAP = 100_000


def _flat_str(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, (dict, list)):
        return json.dumps(v, default=str)[:8000]
    return str(v)


def extract_xccdf_rule_results(raw_xml: bytes) -> list[dict[str, str]]:
    """Parse XCCDF TestResult rule-result elements for spreadsheet export."""
    try:
        from defusedxml import ElementTree

        root = ElementTree.fromstring(raw_xml)
    except Exception:
        return []
    out: list[dict[str, str]] = []
    for el in root.iter():
        tag = el.tag.split("}")[-1] if "}" in (el.tag or "") else (el.tag or "")
        if tag != "rule-result":
            continue
        rule_id = (el.get("idref") or el.get("id") or "").strip()
        sev = (el.get("severity") or "").strip()
        res_text = ""
        for child in el:
            ct = child.tag.split("}")[-1] if "}" in child.tag else child.tag
            if ct == "result" and child.text:
                res_text = (child.text or "").strip()
                break
        out.append(
            {
                "ruleId": rule_id,
                "result": res_text,
                "severity": sev,
            }
        )
    return out


def _is_stig_related_finding(f: dict) -> bool:
    src = (f.get("source") or "").lower()
    if "openscap" in src:
        return True
    if f.get("benchmarkId") or f.get("stableRuleKey"):
        return True
    return False


def _finding_audit_log_rows(findings: list[dict]) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for f in findings:
        fid = _flat_str(f.get("id"))
        for entry in f.get("audit") or []:
            if not isinstance(entry, dict):
                continue
            rows.append(
                {
                    "findingId": fid,
                    "timestamp": _flat_str(entry.get("ts")),
                    "user": _flat_str(entry.get("user")),
                    "action": _flat_str(entry.get("action")),
                    "note": _flat_str(entry.get("note")),
                }
            )
    return rows


def _system_audit_rows(events: list[dict]) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for e in events:
        rows.append(
            {
                "eventId": _flat_str(e.get("event_id")),
                "createdAt": _flat_str(e.get("created_at")),
                "eventType": _flat_str(e.get("event_type")),
                "findingId": _flat_str(e.get("finding_id")),
                "assetId": _flat_str(e.get("asset_id")),
                "sourceId": _flat_str(e.get("source_id")),
                "parserId": _flat_str(e.get("parser_id")),
                "decisionName": _flat_str(e.get("decision_name")),
                "decisionResult": _flat_str(e.get("decision_result")),
            }
        )
    return rows


def _metrics_summary(
    findings: list[dict],
    *,
    generated_at: datetime,
) -> list[dict[str, Any]]:
    open_n = sum(1 for f in findings if is_open_risk(f.get("status") or ""))
    by_sev: dict[str, int] = {}
    for f in findings:
        if not is_open_risk(f.get("status") or ""):
            continue
        sev = f.get("severity") or "Unknown"
        by_sev[sev] = by_sev.get(sev, 0) + 1
    rows = [
        {"metric": "generatedAtUtc", "value": generated_at.isoformat()},
        {"metric": "totalFindingsInExport", "value": len(findings)},
        {"metric": "openFindings", "value": open_n},
    ]
    for sev, c in sorted(by_sev.items(), key=lambda x: x[0]):
        rows.append({"metric": f"openBySeverity_{sev}", "value": c})
    return rows


# ── Style constants ──────────────────────────────────────────────────────────
_HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")  # dark slate
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
_HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
_HEADER_BORDER = Border(
    bottom=Side(style="medium", color="4B5563"),
)
_STRIPE_FILL = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")  # light gray
_CELL_FONT = Font(size=10)
_CELL_ALIGNMENT = Alignment(vertical="top", wrap_text=False)
_CELL_BORDER = Border(
    bottom=Side(style="thin", color="E5E7EB"),
)

_SEVERITY_FILLS: dict[str, PatternFill] = {
    "Critical": PatternFill(start_color="7F1D1D", end_color="7F1D1D", fill_type="solid"),
    "High": PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid"),
    "Medium": PatternFill(start_color="F59E0B", end_color="F59E0B", fill_type="solid"),
    "Low": PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid"),
    "Informational": PatternFill(start_color="6B7280", end_color="6B7280", fill_type="solid"),
}
_SEVERITY_FONTS: dict[str, Font] = {
    "Critical": Font(bold=True, color="FFFFFF", size=10),
    "High": Font(bold=True, color="FFFFFF", size=10),
    "Medium": Font(bold=True, color="000000", size=10),
    "Low": Font(bold=True, color="FFFFFF", size=10),
    "Informational": Font(bold=True, color="FFFFFF", size=10),
}

_STATUS_FILLS: dict[str, PatternFill] = {
    "Open": PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"),
    "In Review": PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid"),
    "Risk Accepted": PatternFill(start_color="E0E7FF", end_color="E0E7FF", fill_type="solid"),
    "Resolved": PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid"),
    "False Positive": PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid"),
}


def _write_sheet(wb: Workbook, title: str, headers: list[str], data_rows: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet(title)

    # Header row
    ws.row_dimensions[1].height = 28
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGNMENT
        cell.border = _HEADER_BORDER

    # Identify severity / status columns for conditional coloring
    sev_col = None
    status_col = None
    for i, h in enumerate(headers):
        hl = h.lower()
        if hl == "severity":
            sev_col = i
        elif hl == "status":
            status_col = i

    # Data rows with alternating stripes and conditional coloring
    for r_i, row in enumerate(data_rows, start=2):
        is_stripe = r_i % 2 == 0
        for c_i, h in enumerate(headers, start=1):
            v = row.get(h)
            if v is None:
                v = ""
            elif isinstance(v, (dict, list)):
                v = str(v)[:32000]
            cell = ws.cell(row=r_i, column=c_i, value=v)
            cell.font = _CELL_FONT
            cell.alignment = _CELL_ALIGNMENT
            cell.border = _CELL_BORDER

            # Severity column: colored badge
            if c_i - 1 == sev_col and isinstance(v, str) and v in _SEVERITY_FILLS:
                cell.fill = _SEVERITY_FILLS[v]
                cell.font = _SEVERITY_FONTS[v]
                cell.alignment = Alignment(horizontal="center", vertical="top")
            # Status column: tinted background
            elif c_i - 1 == status_col and isinstance(v, str) and v in _STATUS_FILLS:
                cell.fill = _STATUS_FILLS[v]
            # Default stripe
            elif is_stripe:
                cell.fill = _STRIPE_FILL

    ws.freeze_panes = "A2"

    # Auto-fit column widths (sample first 50 data rows for performance)
    for i, h in enumerate(headers, start=1):
        max_len = len(h) + 4  # header padding
        col_letter = get_column_letter(i)
        for row in data_rows[:50]:
            val = str(row.get(h) or "")
            max_len = max(max_len, min(len(val) + 2, 60))
        ws.column_dimensions[col_letter].width = min(max(max_len, 12), 60)

    # Auto-filter on header row
    if headers:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(data_rows) + 1}"


def _build_waiver_records_local(rows: list[dict]) -> list[dict]:
    out: list[dict] = []
    for f in rows:
        if (f.get("status") or "") != "Risk Accepted":
            continue
        att = f.get("attestation") if isinstance(f.get("attestation"), dict) else {}
        out.append(
            {
                "findingId": f.get("id"),
                "cveId": f.get("cveId"),
                "ruleId": f.get("ruleId"),
                "title": f.get("title"),
                "severity": f.get("severity"),
                "component": f.get("component"),
                "image": f.get("image"),
                "waiverRef": att.get("waiverRef"),
                "approver": att.get("approver"),
                "approverTitle": att.get("approverTitle"),
                "approvedAt": att.get("approvedAt"),
                "expiresAt": att.get("expiresAt"),
                "controlRef": f.get("controlRef"),
            }
        )
    return out


def _build_readme_sheet(
    ws,
    *,
    generated_at: datetime,
    backend_version: str,
    tenant_id: Optional[str],
    export_options: dict[str, Any],
) -> None:
    """Build a styled cover/ReadMe sheet."""
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 72

    # Title banner (merged across B-C)
    banner_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    banner_font = Font(bold=True, color="FFFFFF", size=16)
    ws.merge_cells("B2:C2")
    title_cell = ws["B2"]
    title_cell.value = "VAT Auditor Workbook"
    title_cell.font = banner_font
    title_cell.fill = banner_fill
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws["C2"].fill = banner_fill
    ws.row_dimensions[2].height = 36

    # Subtitle
    ws.merge_cells("B3:C3")
    sub_cell = ws["B3"]
    sub_cell.value = "Compliance evidence companion to the VAT export bundle"
    sub_cell.font = Font(italic=True, color="6B7280", size=10)

    # Metadata section
    label_font = Font(bold=True, color="374151", size=10)
    value_font = Font(color="111827", size=10)
    section_font = Font(bold=True, color="1F2937", size=11)
    section_border = Border(bottom=Side(style="medium", color="D1D5DB"))

    row = 5
    ws.cell(row=row, column=2, value="Export Details").font = section_font
    ws.cell(row=row, column=2).border = section_border
    ws.cell(row=row, column=3).border = section_border

    metadata = [
        ("Generated (UTC)", generated_at.isoformat().replace("+00:00", "Z")),
        ("Backend Version", backend_version),
        ("Tenant", tenant_id or "(global / unset)"),
        ("Export Options", json.dumps(export_options, default=str)),
    ]
    for label, value in metadata:
        row += 1
        ws.cell(row=row, column=2, value=label).font = label_font
        ws.cell(row=row, column=3, value=value).font = value_font

    # Definitions section
    row += 2
    ws.cell(row=row, column=2, value="Status Definitions").font = section_font
    ws.cell(row=row, column=2).border = section_border
    ws.cell(row=row, column=3).border = section_border

    definitions = [
        ("False Positive", "Scanner incorrect \u2014 global suppression"),
        ("Suppressed", "Real finding accepted in context"),
        ("Risk Accepted", "Formal waiver with attestation"),
    ]
    for label, desc in definitions:
        row += 1
        ws.cell(row=row, column=2, value=label).font = label_font
        ws.cell(row=row, column=3, value=desc).font = value_font

    # Sheet guide section
    row += 2
    ws.cell(row=row, column=2, value="Sheets in this Workbook").font = section_font
    ws.cell(row=row, column=2).border = section_border
    ws.cell(row=row, column=3).border = section_border

    sheets = [
        ("Findings", "All findings with full detail columns"),
        ("Waivers_Risk_Acceptance", "Risk-accepted findings with waiver attestation"),
        ("Justifications_Comments", "Findings with justification or reviewer notes"),
        ("Finding_Decision_Log", "Per-finding audit trail entries"),
        ("Remediation_Backlog", "Open findings requiring action"),
        ("Metrics_Summary", "Aggregate counts and severity breakdown"),
        ("STIG_STIGViewer_Files", "Index of STIG/OpenSCAP XML files in the bundle"),
        ("STIG_Check_Results", "Rule-level results parsed from XCCDF"),
        ("STIG_VAT_Findings", "VAT findings related to STIG benchmarks"),
        ("System_Audit_Events", "System-level audit event stream"),
    ]
    for sheet_name, desc in sheets:
        row += 1
        ws.cell(row=row, column=2, value=sheet_name).font = Font(bold=True, color="2563EB", size=10)
        ws.cell(row=row, column=3, value=desc).font = value_font

    # STIG note
    row += 2
    ws.cell(row=row, column=2, value="STIG Viewer Import").font = section_font
    ws.cell(row=row, column=2).border = section_border
    ws.cell(row=row, column=3).border = section_border
    row += 1
    ws.cell(row=row, column=3, value=(
        "Import raw results from the stig/ folder in the ZIP bundle "
        "(*.xccdf.xml, *.oval-results.xml). See stig/README-STIG-Viewer.txt."
    )).font = value_font
    ws.cell(row=row, column=3).alignment = Alignment(wrap_text=True)
    row += 1
    ws.cell(row=row, column=3, value=(
        f"STIG_Check_Results lists up to {STIG_RULE_ROWS_CAP:,} rule-results parsed from XCCDF; "
        "XML files remain authoritative for STIG Viewer re-import."
    )).font = Font(italic=True, color="6B7280", size=9)
    ws.cell(row=row, column=3).alignment = Alignment(wrap_text=True)

    ws.sheet_properties.tabColor = "1F2937"


def build_auditor_workbook_bytes(
    *,
    findings: list[dict],
    waiver_records: Optional[list[dict]] = None,
    stig_file_manifest: list[dict[str, Any]],
    stig_rule_rows: list[dict[str, Any]],
    audit_events: Optional[list[dict]] = None,
    generated_at: datetime,
    tenant_id: Optional[str],
    backend_version: str,
    export_options: dict[str, Any],
) -> bytes:
    """
    Build auditor-workbook.xlsx: findings, waivers, justifications, decision logs,
    metrics, STIG Viewer file index, STIG check results (from XCCDF), remediation backlog.
    """
    waiver_records = waiver_records if waiver_records is not None else _build_waiver_records_local(findings)

    wb = Workbook()
    default_ws = wb.active
    wb.remove(default_ws)

    # --- ReadMe (styled cover sheet) ---
    ws_readme = wb.create_sheet("ReadMe")
    _build_readme_sheet(
        ws_readme,
        generated_at=generated_at,
        backend_version=backend_version,
        tenant_id=tenant_id,
        export_options=export_options,
    )

    # --- Findings (wide) ---
    finding_headers = [
        "id",
        "fingerprintId",
        "cveId",
        "ruleId",
        "stableRuleKey",
        "benchmarkId",
        "benchmarkFamily",
        "profileScope",
        "title",
        "findingType",
        "severity",
        "status",
        "source",
        "component",
        "image",
        "filePath",
        "team",
        "owner",
        "controlRef",
        "suppressionScope",
        "archived",
        "firstDetectedAt",
        "created",
        "closedAt",
        "slaDue",
        "justification",
        "compensatingControls",
        "reviewerNote",
        "waiverRef",
        "approver",
        "approverTitle",
        "approvedAt",
        "expiresAt",
    ]
    finding_rows: list[dict[str, Any]] = []
    for f in findings:
        att = f.get("attestation") if isinstance(f.get("attestation"), dict) else {}
        finding_rows.append(
            {
                "id": f.get("id"),
                "fingerprintId": f.get("fingerprintId"),
                "cveId": f.get("cveId"),
                "ruleId": f.get("ruleId"),
                "stableRuleKey": f.get("stableRuleKey"),
                "benchmarkId": f.get("benchmarkId"),
                "benchmarkFamily": f.get("benchmarkFamily"),
                "profileScope": f.get("profileScope"),
                "title": f.get("title"),
                "findingType": f.get("findingType"),
                "severity": f.get("severity"),
                "status": f.get("status"),
                "source": f.get("source"),
                "component": f.get("component"),
                "image": f.get("image"),
                "filePath": f.get("filePath"),
                "team": f.get("team"),
                "owner": f.get("owner"),
                "controlRef": f.get("controlRef"),
                "suppressionScope": f.get("suppressionScope"),
                "archived": f.get("archived"),
                "firstDetectedAt": f.get("firstDetectedAt"),
                "created": f.get("created"),
                "closedAt": f.get("closedAt"),
                "slaDue": f.get("slaDue"),
                "justification": f.get("justification"),
                "compensatingControls": f.get("compensatingControls"),
                "reviewerNote": f.get("reviewerNote"),
                "waiverRef": att.get("waiverRef"),
                "approver": att.get("approver"),
                "approverTitle": att.get("approverTitle"),
                "approvedAt": att.get("approvedAt"),
                "expiresAt": att.get("expiresAt"),
            }
        )
    _write_sheet(wb, "Findings", finding_headers, finding_rows)

    # --- Waivers ---
    w_headers = [
        "findingId",
        "cveId",
        "ruleId",
        "title",
        "severity",
        "component",
        "image",
        "waiverRef",
        "approver",
        "approverTitle",
        "approvedAt",
        "expiresAt",
        "controlRef",
    ]
    _write_sheet(wb, "Waivers_Risk_Acceptance", w_headers, waiver_records)

    # --- Justifications (long text) ---
    just_headers = ["findingId", "cveId", "title", "status", "justification", "compensatingControls", "reviewerNote"]
    just_rows = [
        {
            "findingId": f.get("id"),
            "cveId": f.get("cveId"),
            "title": f.get("title"),
            "status": f.get("status"),
            "justification": f.get("justification"),
            "compensatingControls": f.get("compensatingControls"),
            "reviewerNote": f.get("reviewerNote"),
        }
        for f in findings
        if f.get("justification") or f.get("compensatingControls") or f.get("reviewerNote")
    ]
    _write_sheet(wb, "Justifications_Comments", just_headers, just_rows)

    # --- Per-finding audit trail (embedded audit[]) ---
    ad_headers = ["findingId", "timestamp", "user", "action", "note"]
    _write_sheet(wb, "Finding_Decision_Log", ad_headers, _finding_audit_log_rows(findings))

    # --- Remediation-style backlog (not formal eMASS POA&M) ---
    poam_headers = [
        "findingId",
        "cveId",
        "ruleId",
        "title",
        "severity",
        "status",
        "slaDue",
        "team",
        "owner",
        "controlRef",
        "component",
        "image",
        "firstDetectedAt",
    ]
    poam_rows = [
        {
            "findingId": f.get("id"),
            "cveId": f.get("cveId"),
            "ruleId": f.get("ruleId"),
            "title": f.get("title"),
            "severity": f.get("severity"),
            "status": f.get("status"),
            "slaDue": f.get("slaDue"),
            "team": f.get("team"),
            "owner": f.get("owner"),
            "controlRef": f.get("controlRef"),
            "component": f.get("component"),
            "image": f.get("image"),
            "firstDetectedAt": f.get("firstDetectedAt"),
        }
        for f in findings
        if is_open_risk(f.get("status") or "")
    ]
    _write_sheet(wb, "Remediation_Backlog", poam_headers, poam_rows)

    # --- Metrics ---
    _write_sheet(
        wb,
        "Metrics_Summary",
        ["metric", "value"],
        _metrics_summary(findings, generated_at=generated_at),
    )

    # --- STIG files for Viewer / Manager ---
    stig_f_headers = [
        "zipPath",
        "assetId",
        "sourceId",
        "parserId",
        "benchmarkId",
        "benchmarkFamily",
        "profileScope",
        "contentVersion",
        "evidenceSha256",
        "createdAt",
        "importHint",
    ]
    stig_f_rows: list[dict[str, Any]] = []
    for m in stig_file_manifest:
        fn = m.get("filename") or ""
        stig_f_rows.append(
            {
                "zipPath": f"stig/{fn}",
                "assetId": m.get("assetId"),
                "sourceId": m.get("sourceId"),
                "parserId": m.get("parserId"),
                "benchmarkId": m.get("benchmarkId"),
                "benchmarkFamily": m.get("benchmarkFamily"),
                "profileScope": m.get("profileScope"),
                "contentVersion": m.get("contentVersion"),
                "evidenceSha256": m.get("evidenceSha256"),
                "createdAt": m.get("createdAt"),
                "importHint": "DISA STIG Viewer: Import results file (XCCDF or OVAL per tool support).",
            }
        )
    _write_sheet(wb, "STIG_STIGViewer_Files", stig_f_headers, stig_f_rows)

    # --- Rule-level rows from XCCDF ---
    chk_headers = [
        "assetId",
        "sourceId",
        "parserId",
        "benchmarkId",
        "ruleId",
        "result",
        "severity",
    ]
    _write_sheet(wb, "STIG_Check_Results", chk_headers, stig_rule_rows)

    # --- VAT STIG-related findings (normalized) ---
    stig_findings = [f for f in findings if _is_stig_related_finding(f)]
    sf_subset = [
        "id",
        "cveId",
        "ruleId",
        "stableRuleKey",
        "benchmarkId",
        "benchmarkFamily",
        "title",
        "severity",
        "status",
        "source",
        "image",
        "component",
        "filePath",
        "controlRef",
        "suppressionScope",
        "justification",
        "firstDetectedAt",
    ]
    sf_rows = [{k: f.get(k) for k in sf_subset} for f in stig_findings]
    _write_sheet(wb, "STIG_VAT_Findings", sf_subset, sf_rows)

    # --- System audit (same scope as audit-events.json when included) ---
    if audit_events is not None:
        sa_headers = [
            "eventId",
            "createdAt",
            "eventType",
            "findingId",
            "assetId",
            "sourceId",
            "parserId",
            "decisionName",
            "decisionResult",
        ]
        _write_sheet(wb, "System_Audit_Events", sa_headers, _system_audit_rows(audit_events))
    else:
        _write_sheet(
            wb,
            "System_Audit_Events",
            ["note"],
            [
                {
                    "note": "No rows: enable include_audit_events on export bundle to embed audit-events.json and this sheet.",
                }
            ],
        )

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


STIG_VIEWER_README = """DISA STIG Viewer / STIG Manager — importing OpenSCAP results from this ZIP
================================================================================

This folder contains OpenSCAP evaluation results as delivered by VAT:

  *.xccdf.xml       — XCCDF Benchmark result (typical OpenSCAP -xccdf output)
  *.oval-results.xml — OVAL results (when the scan was stored as OVAL results)

HOW TO IMPORT (STIG Viewer)
---------------------------
1. Open DISA STIG Viewer.
2. Use the menu action to import or open **result** / **XCCDF** data (wording varies by version).
3. Select one or more `.xccdf.xml` files from this `stig/` folder.
4. Repeat per asset or scan file as needed.

OVAL-only files may be supported depending on your Viewer version; if import fails,
use the corresponding XCCDF result from the same scan when available.

AUTHORITATIVE DATA
------------------
The XML files in this folder are the authoritative machine format for re-import.
`stig/manifest.json` lists files and metadata. The Excel workbook
`auditor-workbook.xlsx` (in the bundle root) includes sheets
`STIG_STIGViewer_Files`, `STIG_Check_Results`, and `STIG_VAT_Findings` for human review.

For questions on VAT export layout, see `evidence-manifest.json` in the bundle root.
""".strip()
