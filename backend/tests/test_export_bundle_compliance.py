"""Compliance export bundle: manifest, CSV, waivers, PDF, audit embed."""

from __future__ import annotations

import csv
from datetime import datetime, timezone
import hashlib
import io
import json
import zipfile
from types import SimpleNamespace
from unittest.mock import AsyncMock, MagicMock

import pytest

from app.models.finding import Finding, FindingType, Severity, Status
from app.services.export_service import (
    ExportBundleOptions,
    _build_waiver_records,
    _finding_csv_row,
    build_export_bundle,
)
from app.services.audit_workbook_export import _iter_finding_workbook_rows


def _waiver_finding(finding_id: str = "wf1", image: str = "api:latest") -> Finding:
    ts = datetime(2025, 1, 10, 12, 0, 0, tzinfo=timezone.utc)
    return Finding(
        id=finding_id,
        finding_type=FindingType.SCA,
        fingerprint_id="fp_waiver_1",
        cve_id="CVE-2024-99999",
        severity=Severity.High,
        status=Status.RiskAccepted,
        title="Test waiver row",
        component="openssl 3",
        image=image,
        control_ref="CC6.6",
        attestation={
            "approver": "Sam Reviewer",
            "approverTitle": "CISO",
            "approvedAt": "2025-01-15T10:00:00Z",
            "waiverRef": "WAV-TEST-001",
            "expiresAt": "2026-12-31",
        },
        sources=[],
        external_links=[],
        tracker_comment=False,
        regression_count=0,
        audit=[],
        archived=False,
        created_at=ts,
        updated_at=ts,
        first_detected_at=ts,
    )


def test_finding_csv_row_includes_risk_scoring_columns():
    row = _finding_csv_row(
        {
            "id": "f-risk",
            "riskScoring": {
                "source": {
                    "score": "7.4",
                    "severity": "High",
                    "vector": "CVSS:3.1/AV:N/AC:H/PR:N/UI:N/S:U/C:H/I:H/A:N",
                    "fixedVersion": "NONE",
                },
                "threat": {"epss": "0.012", "knownExploited": False},
                "context": {"reachability": "No path found", "fixAvailable": False},
                "environmental": {
                    "score": "0.0",
                    "vector": "CVSS:3.1/AV:N/AC:H/PR:N/UI:N/S:U/C:H/I:H/A:N/MC:N/MI:N/MA:N",
                    "rationale": "Vulnerable ECDSA path is not reachable.",
                },
            },
        }
    )

    assert row["sourceCvssScore"] == "7.4"
    assert row["sourceCvssSeverity"] == "High"
    assert row["fixedVersion"] == "NONE"
    assert row["epss"] == "0.012"
    assert row["knownExploited"] == "False"
    assert row["reachability"] == "No path found"
    assert row["fixAvailable"] == "False"
    assert row["environmentalScore"] == "0.0"
    assert row["environmentalRationale"] == "Vulnerable ECDSA path is not reachable."


def test_auditor_workbook_findings_include_risk_scoring_columns():
    row = next(
        _iter_finding_workbook_rows(
            [
                {
                    "id": "f-risk",
                    "riskScoring": {
                        "source": {"score": "7.4", "fixedVersion": "NONE"},
                        "environmental": {
                            "score": "0.0",
                            "rationale": "Not reachable.",
                        },
                    },
                }
            ]
        )
    )

    assert row["sourceCvssScore"] == "7.4"
    assert row["fixedVersion"] == "NONE"
    assert row["environmentalScore"] == "0.0"
    assert row["environmentalRationale"] == "Not reachable."


@pytest.mark.asyncio
async def test_build_export_bundle_compliance_artifacts_and_manifest_hashes(monkeypatch):
    async def mock_list_findings(*_a, **_k):
        return [_waiver_finding()]

    monkeypatch.setattr(
        "app.services.export_service.list_findings",
        mock_list_findings,
    )
    monkeypatch.setattr(
        "app.services.export_service.enrich_findings_with_source_group_severity",
        AsyncMock(side_effect=lambda _db, rows: rows),
    )
    monkeypatch.setattr(
        "app.services.export_service.get_assets_with_findings",
        AsyncMock(return_value=[]),
    )
    monkeypatch.setattr(
        "app.services.export_service.list_sbom_packages",
        AsyncMock(return_value=[]),
    )
    monkeypatch.setattr(
        "app.services.export_service.list_openscap_scan_results",
        AsyncMock(return_value=[]),
    )
    monkeypatch.setattr(
        "app.services.export_service.load_audit_events_for_export",
        AsyncMock(
            return_value=[
                {"event_id": "evt-1", "event_type": "test.event", "data": {}},
            ]
        ),
    )

    db = MagicMock()
    ctx = SimpleNamespace(tenant_id="tenant-a", cross_tenant=False)
    data = await build_export_bundle(
        db,
        ctx=ctx,
        options=ExportBundleOptions(include_audit_events=True),
    )
    assert data.startswith(b"PK")

    zf = zipfile.ZipFile(io.BytesIO(data), "r")
    names = zf.namelist()
    prefix = [n for n in names if n.startswith("vat-export-")][0].split("/")[0]

    expected = {
        f"{prefix}/evidence-manifest.json",
        f"{prefix}/assets-findings.json",
        f"{prefix}/findings.csv",
        f"{prefix}/waivers.json",
        f"{prefix}/waivers.csv",
        f"{prefix}/executive-summary-yearly.html",
        f"{prefix}/audit-events.json",
        f"{prefix}/sbom/sbom-cyclonedx.json",
        f"{prefix}/auditor-workbook.xlsx",
        f"{prefix}/stig/README-STIG-Viewer.txt",
    }
    for path in expected:
        assert path in names, f"missing {path}"
    assert f"{prefix}/compliance-summary.pdf" not in names

    manifest = json.loads(zf.read(f"{prefix}/evidence-manifest.json").decode())
    assert manifest["schemaVersion"] == "evidence-v2"
    assert manifest["packageType"] == "vat-compliance-bundle"
    assert manifest["tenantId"] == "tenant-a"
    assert "vatBackendVersion" in manifest
    assert manifest["exportOptions"]["include_audit_events"] is True

    by_path = {e["path"]: e for e in manifest["files"]}
    for rel, zpath in [
        ("findings.csv", f"{prefix}/findings.csv"),
        ("audit-events.json", f"{prefix}/audit-events.json"),
    ]:
        raw = zf.read(zpath)
        assert by_path[rel]["sha256"] == hashlib.sha256(raw).hexdigest()
        assert by_path[rel]["sizeBytes"] == len(raw)

    waivers = json.loads(zf.read(f"{prefix}/waivers.json").decode())
    assert len(waivers) == 1
    assert waivers[0]["waiverRef"] == "WAV-TEST-001"
    assert waivers[0]["approver"] == "Sam Reviewer"

    wcsv = zf.read(f"{prefix}/waivers.csv").decode()
    r = list(csv.DictReader(io.StringIO(wcsv)))
    assert len(r) == 1
    assert r[0]["waiverRef"] == "WAV-TEST-001"


@pytest.mark.asyncio
async def test_build_export_bundle_uses_id_only_asset_payloads(monkeypatch):
    async def mock_list_findings(*_a, **_k):
        return [_waiver_finding("wf1", "asset-a")]

    get_assets_spy = AsyncMock(
        return_value=[
            {
                "id": "asset-a",
                "name": "asset-a",
                "findingIds": ["wf1"],
                "findings": [],
            }
        ]
    )
    monkeypatch.setattr("app.services.export_service.list_findings", mock_list_findings)
    monkeypatch.setattr(
        "app.services.export_service.enrich_findings_with_source_group_severity",
        AsyncMock(side_effect=lambda _db, rows: rows),
    )
    monkeypatch.setattr(
        "app.services.export_service.get_assets_with_findings", get_assets_spy
    )
    monkeypatch.setattr(
        "app.services.export_service.list_sbom_packages", AsyncMock(return_value=[])
    )
    monkeypatch.setattr(
        "app.services.export_service.list_openscap_scan_results",
        AsyncMock(return_value=[]),
    )
    monkeypatch.setattr(
        "app.services.export_service.load_audit_events_for_export",
        AsyncMock(return_value=[]),
    )

    db = MagicMock()
    ctx = SimpleNamespace(tenant_id="tenant-a", cross_tenant=False)
    data = await build_export_bundle(
        db, ctx=ctx, options=ExportBundleOptions(include_audit_events=False)
    )

    get_assets_spy.assert_awaited_once()
    assert get_assets_spy.await_args.kwargs["include_findings"] is False

    zf = zipfile.ZipFile(io.BytesIO(data), "r")
    prefix = [n for n in zf.namelist() if n.startswith("vat-export-")][0].split("/")[0]
    payload = json.loads(zf.read(f"{prefix}/assets-findings.json").decode())
    assert payload["assets"][0]["findingIds"] == ["wf1"]
    assert payload["assets"][0]["findings"] == []


@pytest.mark.asyncio
async def test_build_export_bundle_skips_audit_when_disabled(monkeypatch):
    async def mock_list_findings(*_a, **_k):
        return []

    monkeypatch.setattr(
        "app.services.export_service.list_findings",
        mock_list_findings,
    )
    monkeypatch.setattr(
        "app.services.export_service.enrich_findings_with_source_group_severity",
        AsyncMock(side_effect=lambda _db, rows: rows),
    )
    monkeypatch.setattr(
        "app.services.export_service.get_assets_with_findings",
        AsyncMock(return_value=[]),
    )
    monkeypatch.setattr(
        "app.services.export_service.list_sbom_packages",
        AsyncMock(return_value=[]),
    )
    monkeypatch.setattr(
        "app.services.export_service.list_openscap_scan_results",
        AsyncMock(return_value=[]),
    )
    spy = AsyncMock(return_value=[])
    monkeypatch.setattr(
        "app.services.export_service.load_audit_events_for_export",
        spy,
    )

    db = MagicMock()
    ctx = SimpleNamespace(tenant_id="tenant-a", cross_tenant=False)
    await build_export_bundle(
        db,
        ctx=ctx,
        options=ExportBundleOptions(include_audit_events=False),
    )
    spy.assert_not_called()


@pytest.mark.asyncio
async def test_build_export_bundle_scopes_to_selected_assets(monkeypatch):
    async def mock_list_findings(*_a, **_k):
        return [
            _waiver_finding("wf1", "asset-a"),
            _waiver_finding("wf2", "asset-b"),
        ]

    monkeypatch.setattr(
        "app.services.export_service.list_findings",
        mock_list_findings,
    )
    monkeypatch.setattr(
        "app.services.export_service.enrich_findings_with_source_group_severity",
        AsyncMock(side_effect=lambda _db, rows: rows),
    )
    monkeypatch.setattr(
        "app.services.export_service.get_assets_with_findings",
        AsyncMock(
            return_value=[
                {"id": "asset-a", "name": "asset-a", "findingIds": ["wf1"], "findings": []},
                {"id": "asset-b", "name": "asset-b", "findingIds": ["wf2"], "findings": []},
            ]
        ),
    )
    monkeypatch.setattr(
        "app.services.export_service.list_sbom_packages",
        AsyncMock(return_value=[]),
    )
    monkeypatch.setattr(
        "app.services.export_service.list_openscap_scan_results",
        AsyncMock(return_value=[]),
    )
    monkeypatch.setattr(
        "app.services.export_service.load_audit_events_for_export",
        AsyncMock(return_value=[]),
    )

    db = MagicMock()
    ctx = SimpleNamespace(tenant_id="tenant-a", cross_tenant=False)
    data = await build_export_bundle(
        db,
        ctx=ctx,
        options=ExportBundleOptions(
            apply_asset_filter=True,
            asset_ids=["asset-a"],
            include_audit_events=False,
        ),
    )
    zf = zipfile.ZipFile(io.BytesIO(data), "r")
    names = zf.namelist()
    prefix = [n for n in names if n.startswith("vat-export-")][0].split("/")[0]
    payload = json.loads(zf.read(f"{prefix}/assets-findings.json").decode())
    assert [f["id"] for f in payload["findings"]] == ["wf1"]
    assert [a["id"] for a in payload["assets"]] == ["asset-a"]


@pytest.mark.asyncio
async def test_build_export_bundle_includes_scan_evidence_artifacts(monkeypatch):
    xccdf_xml = b"""<?xml version="1.0" encoding="UTF-8"?>
<Benchmark xmlns="http://checklists.nist.gov/xccdf/1.2" id="xccdf_org.ssgproject.content_benchmark_TEST">
  <TestResult id="xccdf_org.open-scap_testresult_default-profile" profile="stig">
    <rule-result idref="xccdf_org.ssgproject.content_rule_file_permissions" severity="medium">
      <result>fail</result>
    </rule-result>
  </TestResult>
</Benchmark>
"""
    openscap_row = SimpleNamespace(
        asset_id="k8s/cluster/node/node-a/runtime-image/app-1.0",
        source_id="openscap",
        parser_id="openscap",
        raw_xccdf_xml=xccdf_xml,
        created_at=datetime(2025, 6, 1, tzinfo=timezone.utc),
        benchmark_id="xccdf_org.ssgproject.content_benchmark_TEST",
        benchmark_family="TEST",
        profile_scope="stig",
        content_version="1",
        evidence_sha256="abc123",
    )

    monkeypatch.setattr("app.services.export_service.list_findings", AsyncMock(return_value=[]))
    monkeypatch.setattr(
        "app.services.export_service.enrich_findings_with_source_group_severity",
        AsyncMock(side_effect=lambda _db, rows: rows),
    )
    monkeypatch.setattr(
        "app.services.export_service.get_assets_with_findings",
        AsyncMock(return_value=[]),
    )
    monkeypatch.setattr(
        "app.services.export_service.list_sbom_packages",
        AsyncMock(
            return_value=[
                {
                    "id": "pkg-1",
                    "name": "openssl",
                    "version": "3.0.0",
                    "licenseId": "Apache-2.0",
                    "component": "ghcr.io/acme/app:1.0",
                    "language": "c",
                }
            ]
        ),
    )
    monkeypatch.setattr(
        "app.services.export_service.list_openscap_scan_results",
        AsyncMock(return_value=[openscap_row]),
    )
    monkeypatch.setattr(
        "app.services.export_service.load_audit_events_for_export",
        AsyncMock(return_value=[]),
    )

    data = await build_export_bundle(
        MagicMock(),
        ctx=SimpleNamespace(tenant_id="tenant-a", cross_tenant=False),
        options=ExportBundleOptions(include_audit_events=True),
    )

    zf = zipfile.ZipFile(io.BytesIO(data), "r")
    names = zf.namelist()
    prefix = [n for n in names if n.startswith("vat-export-")][0].split("/")[0]
    stig_files = [n for n in names if n.startswith(f"{prefix}/stig/")]
    assert f"{prefix}/stig/README-STIG-Viewer.txt" in names
    assert f"{prefix}/stig/manifest.json" in names
    assert any(n.endswith(".xccdf.xml") for n in stig_files)
    assert f"{prefix}/sbom/sbom-cyclonedx.json" in names
    assert f"{prefix}/sbom/by-asset/manifest.json" in names
    assert f"{prefix}/auditor-workbook.xlsx" in names

    stig_manifest = json.loads(zf.read(f"{prefix}/stig/manifest.json").decode())
    assert stig_manifest[0]["assetId"] == openscap_row.asset_id
    assert stig_manifest[0]["parserId"] == "openscap"
    asset_manifest = json.loads(zf.read(f"{prefix}/sbom/by-asset/manifest.json").decode())
    assert asset_manifest[0]["component"] == "ghcr.io/acme/app:1.0"

    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(zf.read(f"{prefix}/auditor-workbook.xlsx")), read_only=True)
    assert "STIG_STIGViewer_Files" in wb.sheetnames
    assert "STIG_Check_Results" in wb.sheetnames
    check_rows = list(wb["STIG_Check_Results"].iter_rows(values_only=True))
    assert check_rows[1][4] == "xccdf_org.ssgproject.content_rule_file_permissions"
    assert check_rows[1][5] == "fail"


def test_build_waiver_records_filters_status():
    rows = [
        {"id": "a", "status": "Open", "attestation": {}},
        {"id": "b", "status": "Risk Accepted", "attestation": {"waiverRef": "W"}},
    ]
    rec = _build_waiver_records(rows)
    assert len(rec) == 1
    assert rec[0]["findingId"] == "b"
